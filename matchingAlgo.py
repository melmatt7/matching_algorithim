from openpyxl import load_workbook

mentee_invalid_mentee_id = []
mentee_invalid_mentor_id = []

mentor_invalid_mentee_id = []
mentor_invalid_mentor_id = []
mentor_invalid_mentor_count = []

# restrictions need to be tweaked
MAX_MENTEE_ID = 748000
MIN_MENTEE_ID = 747000

MAX_MENTOR_ID = 150
MIN_MENTOR_ID = 0

MAX_COUNT = 10
MIN_COUNT = 0

MENTEE_APPLICATIONS = 386
MENTOR_APPLICATIONS = 101

def matching():
	mentor_book = load_workbook("2018-2019 REX_ Applicant Ranking.xlsx")
	mentee_book = load_workbook("2018 - 2019 REX Mentee Application.xlsx")

	mentee = mentee_book.active
	mentor = mentor_book.active

	mentee_info = parse_mentee(mentee)
	mentor_info = parse_mentor(mentor)

	print(mentor_info)
	print(mentor_invalid_mentee_id)
	print(mentor_invalid_mentor_id)

def parse_mentee(mentee):

	mentee_id = []
	mentee_1 = []
	mentee_2 = []
	mentee_3 = []
	mentee_4 = []
	mentee_5 = []
	mentee_info = []

	for i in range(2,MENTEE_APPLICATIONS): 

		mentee_id_val = mentee.cell(row=i, column=6).value # column 6 is F 
		parseVal(None, mentee_id_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentee_id, mentee_invalid_mentee_id)
		
		mentee_1_val = mentee.cell(row=i, column=15).value # column 15 is O 
		parseVal(mentee_id_val, mentee_1_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentee_1, mentee_invalid_mentor_id)

		mentee_2_val = mentee.cell(row=i, column=18).value # column 18 is R 
		parseVal(mentee_id_val, mentee_2_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentee_2, mentee_invalid_mentor_id)

		mentee_3_val = mentee.cell(row=i, column=21).value # column 21 is U 
		parseVal(mentee_id_val, mentee_3_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentee_3, mentee_invalid_mentor_id)

		mentee_4_val = mentee.cell(row=i, column=24).value # column 24 is X 
		parseVal(mentee_id_val, mentee_4_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentee_4, mentee_invalid_mentor_id)

		mentee_5_val = mentee.cell(row=i, column=27).value # column 27 is AA 
		parseVal(mentee_id_val, mentee_5_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentee_5, mentee_invalid_mentor_id)	

		mentee_info.append([mentee_id[-1], mentee_1[-1], mentee_2[-1], mentee_3[-1], mentee_4[-1], mentee_5[-1]])

	return mentee_info

def parse_mentor(mentor):
	mentor_id = []
	mentor_count = []
	mentor_1 = []
	mentor_2 = []
	mentor_3 = []
	mentor_4 = []
	mentor_5 = []
	mentor_6 = []
	mentor_7 = []
	mentor_8 = []
	mentor_9 = []
	mentor_10 = []
	mentor_11 = []
	mentor_12 = []
	mentor_13 = []
	mentor_14 = []
	mentor_15 = []
	mentor_16 = []
	mentor_17 = []
	mentor_18 = []
	mentor_19 = []
	mentor_20 = []
	mentor_info = []

	for i in range(2,MENTOR_APPLICATIONS): 

		mentor_id_val = mentor.cell(row=i, column=3).value # column 3 is C 
		parseVal(None, mentor_id_val, MAX_MENTOR_ID, MIN_MENTOR_ID, mentor_id, mentor_invalid_mentor_id)

		mentor_count_val = mentor.cell(row=i, column=4).value # column 4 is D 
		parseVal(mentor_id_val, mentor_count_val, MAX_COUNT, MIN_COUNT, mentor_count, mentor_invalid_mentor_count)
		
		mentor_1_val = mentor.cell(row=i, column=6).value # column 6 is F 
		parseVal(mentor_id_val, mentor_1_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_1, mentor_invalid_mentee_id)

		mentor_2_val = mentor.cell(row=i, column=8).value # column 8 is H 
		parseVal(mentor_id_val, mentor_2_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_2, mentor_invalid_mentee_id)

		mentor_3_val = mentor.cell(row=i, column=10).value # column 10 is J 
		parseVal(mentor_id_val, mentor_3_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_3, mentor_invalid_mentee_id)

		mentor_4_val = mentor.cell(row=i, column=12).value # column 12 is L 
		parseVal(mentor_id_val, mentor_4_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_4, mentor_invalid_mentee_id)

		mentor_5_val = mentor.cell(row=i, column=14).value # column 14 is N 
		parseVal(mentor_id_val, mentor_5_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_5, mentor_invalid_mentee_id)

		mentor_6_val = mentor.cell(row=i, column=16).value # column 16 is P 
		parseVal(mentor_id_val, mentor_6_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_6, mentor_invalid_mentee_id)	

		mentor_7_val = mentor.cell(row=i, column=18).value # column 18 is R 
		parseVal(mentor_id_val, mentor_7_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_7, mentor_invalid_mentee_id)

		mentor_8_val = mentor.cell(row=i, column=20).value # column 20 is T
		parseVal(mentor_id_val, mentor_8_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_8, mentor_invalid_mentee_id)

		mentor_9_val = mentor.cell(row=i, column=22).value # column 22 is V 
		parseVal(mentor_id_val, mentor_9_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_9, mentor_invalid_mentee_id)

		mentor_10_val = mentor.cell(row=i, column=24).value # column 24 is X 
		parseVal(mentor_id_val, mentor_10_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_10, mentor_invalid_mentee_id)

		mentor_11_val = mentor.cell(row=i, column=26).value # column 26 is Z 
		parseVal(mentor_id_val, mentor_11_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_11, mentor_invalid_mentee_id)

		mentor_12_val = mentor.cell(row=i, column=28).value # column 28 is AB 
		parseVal(mentor_id_val, mentor_12_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_12, mentor_invalid_mentee_id)

		mentor_13_val = mentor.cell(row=i, column=30).value # column 30 is AD 
		parseVal(mentor_id_val, mentor_13_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_13, mentor_invalid_mentee_id)

		mentor_14_val = mentor.cell(row=i, column=32).value # column 32 is AF 
		parseVal(mentor_id_val, mentor_14_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_14, mentor_invalid_mentee_id)

		mentor_15_val = mentor.cell(row=i, column=34).value # column 34 is AH 
		parseVal(mentor_id_val, mentor_15_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_15, mentor_invalid_mentee_id)

		mentor_16_val = mentor.cell(row=i, column=36).value # column 36 is AJ 
		parseVal(mentor_id_val, mentor_16_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_16, mentor_invalid_mentee_id)

		mentor_17_val = mentor.cell(row=i, column=38).value # column 38 is AL 
		parseVal(mentor_id_val, mentor_17_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_17, mentor_invalid_mentee_id)

		mentor_18_val = mentor.cell(row=i, column=40).value # column 40 is AN 
		parseVal(mentor_id_val, mentor_18_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_18, mentor_invalid_mentee_id)

		mentor_19_val = mentor.cell(row=i, column=42).value # column 42 is AP 
		parseVal(mentor_id_val, mentor_19_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_19, mentor_invalid_mentee_id)

		mentor_20_val = mentor.cell(row=i, column=44).value # column 44 is AR 
		parseVal(mentor_id_val, mentor_20_val, MAX_MENTEE_ID, MIN_MENTEE_ID, mentor_20, mentor_invalid_mentee_id)

		mentor_info.append([mentor_id[-1], mentor_count[-1], mentor_1[-1], mentor_2[-1], mentor_3[-1], mentor_4[-1], mentor_5[-1], mentor_6[-1], mentor_7[-1], mentor_8[-1], mentor_9[-1], mentor_10[-1], mentor_11[-1], mentor_12[-1], mentor_13[-1], mentor_14[-1], mentor_15[-1], mentor_16[-1], mentor_17[-1], mentor_18[-1], mentor_19[-1], mentor_20[-1]])

	return mentor_info


def parseVal(id, val, max, min, arr, invalid):
	val, flag = checkValid(val, max, min)
	if flag:
		if id == None:
			invalid.append(val)
			arr.append(None)
		elif val == None:
			arr.append(val)
		else:
			invalid.append([id,val])
			arr.append(None)		
	else:
		arr.append(val)

def checkValid(value, max, min):
	if type(value) == float and value >= min and value <= max: 
		return int(value), False
	else:
		if type(value) == float:
			return int(value), True
		return value, True

matching()